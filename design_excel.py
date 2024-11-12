from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference

# Create a new Excel workbook and add a worksheet
workbook = Workbook()
sheet = workbook.active
sheet.title = "Test"

# Remove gridlines
sheet.sheet_view.showGridLines = False

# Define the thick border style
thick_border = Border(left=Side(style='thick', color='000000'),
                      right=Side(style='thick', color='000000'),
                      top=Side(style='thick', color='000000'),
                      bottom=Side(style='thick', color='000000'))

# Remove all existing borders in the range D5 to Y62
for row in range(5, 63):
    for col in range(4, 26):  # D is the 4th column, Y is the 25th column
        cell = sheet.cell(row=row, column=col)
        cell.border = Border()

# Apply the thick border around the range D5 to Y62
for row in range(5, 63):
    # Left border
    sheet.cell(row=row, column=4).border = Border(left=thick_border.left)
    # Right border
    sheet.cell(row=row, column=25).border = Border(right=thick_border.right)

for col in range(4, 26):
    # Top border
    sheet.cell(row=5, column=col).border = Border(top=thick_border.top)
    # Bottom border
    sheet.cell(row=62, column=col).border = Border(bottom=thick_border.bottom)

# Set the corners
sheet.cell(row=5, column=4).border = Border(top=thick_border.top, left=thick_border.left)
sheet.cell(row=5, column=25).border = Border(top=thick_border.top, right=thick_border.right)
sheet.cell(row=62, column=4).border = Border(bottom=thick_border.bottom, left=thick_border.left)
sheet.cell(row=62, column=25).border = Border(bottom=thick_border.bottom, right=thick_border.right)

# Set the text "hubway-data" in a merged cell
sheet.merge_cells('D5:Y6')
cell = sheet.cell(row=5, column=4)
cell.value = "hubway-data"
cell.font = Font(color="FFFFFF")  # White font
cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Gray background
cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

# Merge cells E8 to H13 and set the text "February"
sheet.merge_cells('E8:H13')
cell = sheet.cell(row=8, column=5)
cell.value = "February"
cell.font = Font(color="000000")  # Black font
cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

# Apply thick border around the "February" box
for row in range(8, 14):
    # Left border
    sheet.cell(row=row, column=5).border = Border(left=thick_border.left)
    # Right border
    sheet.cell(row=row, column=8).border = Border(right=thick_border.right)

for col in range(5, 9):
    # Top border
    sheet.cell(row=8, column=col).border = Border(top=thick_border.top)
    # Bottom border
    sheet.cell(row=13, column=col).border = Border(bottom=thick_border.bottom)

# Set the corners for the "February" box
sheet.cell(row=8, column=5).border = Border(top=thick_border.top, left=thick_border.left)
sheet.cell(row=8, column=8).border = Border(top=thick_border.top, right=thick_border.right)
sheet.cell(row=13, column=5).border = Border(bottom=thick_border.bottom, left=thick_border.left)
sheet.cell(row=13, column=8).border = Border(bottom=thick_border.bottom, right=thick_border.right)

# Add the first table from J15 to M37 with columns "Bike ID" and "Anzahl Nutzung"
sheet.merge_cells('J11:M12')
cell = sheet.cell(row=11, column=10, value="Bike ID and Anzahl Nutzung")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

sheet.merge_cells('J15:K16')
sheet.merge_cells('L15:M16')
cell = sheet.cell(row=15, column=10, value="Bike ID")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell = sheet.cell(row=15, column=12, value="Anzahl Nutzung")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Add 10 rows of sample data for the first table
sample_data = [
    ("Bike001", 15),
    ("Bike002", 20),
    ("Bike003", 25),
    ("Bike004", 30),
    ("Bike005", 35),
    ("Bike006", 40),
    ("Bike007", 45),
    ("Bike008", 50),
    ("Bike009", 55),
    ("Bike010", 60),
]

for i, (bike_id, usage_count) in enumerate(sample_data, start=1):
    row_start = 15 + 2 * i
    row_end = row_start + 1
    sheet.merge_cells(start_row=row_start, start_column=10, end_row=row_end, end_column=11)
    sheet.merge_cells(start_row=row_start, start_column=12, end_row=row_end, end_column=13)
    sheet.cell(row=row_start, column=10, value=bike_id).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row_start, column=12, value=usage_count).alignment = Alignment(horizontal="center", vertical="center")

# Apply thick border around the first table
for row in range(15, 37):
    for col in range(10, 14):
        cell = sheet.cell(row=row, column=col)
        cell.border = thick_border

# Add the second table from O15 to R37 with columns "Top 10 most start stations" and "Anzahl Nutzung"
sheet.merge_cells('O11:R12')
cell = sheet.cell(row=11, column=15, value="Top 10 most start stations and Anzahl Nutzung")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

sheet.merge_cells('O15:P16')
sheet.merge_cells('Q15:R16')
cell = sheet.cell(row=15, column=15, value="Top 10 most start stations")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell = sheet.cell(row=15, column=17, value="Anzahl Nutzung")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Add 10 rows of sample data for the second table
sample_data_start_stations = [
    ("Station001", 150),
    ("Station002", 200),
    ("Station003", 250),
    ("Station004", 300),
    ("Station005", 350),
    ("Station006", 400),
    ("Station007", 450),
    ("Station008", 500),
    ("Station009", 550),
    ("Station010", 600),
]

for i, (station_id, usage_count) in enumerate(sample_data_start_stations, start=1):
    row_start = 15 + 2 * i
    row_end = row_start + 1
    sheet.merge_cells(start_row=row_start, start_column=15, end_row=row_end, end_column=16)
    sheet.merge_cells(start_row=row_start, start_column=17, end_row=row_end, end_column=18)
    sheet.cell(row=row_start, column=15, value=station_id).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row_start, column=17, value=usage_count).alignment = Alignment(horizontal="center", vertical="center")

# Apply thick border around the second table
for row in range(15, 37):
    for col in range(15, 19):
        cell = sheet.cell(row=row, column=col)
        cell.border = thick_border

# Add the third table from T15 to W37 with columns "Top 10 most end stations" and "Anzahl Nutzung"
sheet.merge_cells('T11:W12')
cell = sheet.cell(row=11, column=20, value="Top 10 most end stations and Anzahl Nutzung")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

sheet.merge_cells('T15:U16')
sheet.merge_cells('V15:W16')
cell = sheet.cell(row=15, column=20, value="Top 10 most end stations")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell = sheet.cell(row=15, column=22, value="Anzahl Nutzung")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Add 10 rows of sample data for the third table
sample_data_end_stations = [
    ("Station011", 160),
    ("Station012", 210),
    ("Station013", 260),
    ("Station014", 310),
    ("Station015", 360),
    ("Station016", 410),
    ("Station017", 460),
    ("Station018", 510),
    ("Station019", 560),
    ("Station020", 610),
]

for i, (station_id, usage_count) in enumerate(sample_data_end_stations, start=1):
    row_start = 15 + 2 * i
    row_end = row_start + 1
    sheet.merge_cells(start_row=row_start, start_column=20, end_row=row_end, end_column=21)
    sheet.merge_cells(start_row=row_start, start_column=22, end_row=row_end, end_column=23)
    sheet.cell(row=row_start, column=20, value=station_id).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row_start, column=22, value=usage_count).alignment = Alignment(horizontal="center", vertical="center")

# Apply thick border around the third table
for row in range(15, 37):
    for col in range(20, 24):
        cell = sheet.cell(row=row, column=col)
        cell.border = thick_border

# Add the "Average Trip Duration" header
sheet.merge_cells('E15:H18')
cell = sheet.cell(row=15, column=5, value="Average Trip Duration")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Apply thick border around the "Average Trip Duration" box
for row in range(15, 19):
    # Left border
    sheet.cell(row=row, column=5).border = Border(left=thick_border.left)
    # Right border
    sheet.cell(row=row, column=8).border = Border(right=thick_border.right)

for col in range(5, 9):
    # Top border
    sheet.cell(row=15, column=col).border = Border(top=thick_border.top)
    # Bottom border
    sheet.cell(row=18, column=col).border = Border(bottom=thick_border.bottom)

# Set the corners for the "Average Trip Duration" box
sheet.cell(row=15, column=5).border = Border(top=thick_border.top, left=thick_border.left)
sheet.cell(row=15, column=8).border = Border(top=thick_border.top, right=thick_border.right)
sheet.cell(row=18, column=5).border = Border(bottom=thick_border.bottom, left=thick_border.left)
sheet.cell(row=18, column=8).border = Border(bottom=thick_border.bottom, right=thick_border.right)

# Add sample data for average trip duration
average_trip_duration = "15 minutes"
sheet.merge_cells('E19:H24')
cell = sheet.cell(row=19, column=5, value=average_trip_duration)
cell.alignment = Alignment(horizontal="center", vertical="center")

# Add the "Average Distance" header
sheet.merge_cells('E26:H29')
cell = sheet.cell(row=26, column=5, value="Average Distance")
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Apply thick border around the "Average Distance" box
for row in range(26, 30):
    # Left border
    sheet.cell(row=row, column=5).border = Border(left=thick_border.left)
    # Right border
    sheet.cell(row=row, column=8).border = Border(right=thick_border.right)

for col in range(5, 9):
    # Top border
    sheet.cell(row=26, column=col).border = Border(top=thick_border.top)
    # Bottom border
    sheet.cell(row=29, column=col).border = Border(bottom=thick_border.bottom)

# Set the corners for the "Average Distance" box
sheet.cell(row=26, column=5).border = Border(top=thick_border.top, left=thick_border.left)
sheet.cell(row=26, column=8).border = Border(top=thick_border.top, right=thick_border.right)
sheet.cell(row=29, column=5).border = Border(bottom=thick_border.bottom, left=thick_border.left)
sheet.cell(row=29, column=8).border = Border(bottom=thick_border.bottom, right=thick_border.right)

# Add sample data for average distance
average_distance = "5 km"
sheet.merge_cells('E30:H35')
cell = sheet.cell(row=30, column=5, value=average_distance)
cell.alignment = Alignment(horizontal="center", vertical="center")

# Add sample data for the pie chart
data = [
    ["Category", "Value"],
    ["Category 1", 30],
    ["Category 2", 45],
    ["Category 3", 25],
]

# Add the data to the sheet starting from a specific row
start_row = 38
for i, row in enumerate(data, start=start_row):
    for j, value in enumerate(row, start=5):  # Start from column E (5th column)
        sheet.cell(row=i, column=j, value=value)

# Create a pie chart
pie = PieChart()
labels = Reference(sheet, min_col=5, min_row=start_row + 1, max_row=start_row + 3)
data = Reference(sheet, min_col=6, min_row=start_row + 1, max_row=start_row + 3)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Sample Pie Chart"

# Set the size of the pie chart
pie.width = 10  # Width in columns (E to K)
pie.height = 12  # Height in rows (38 to 56)

# Add the pie chart to the sheet
sheet.add_chart(pie, "E38")

# Add sample data for the second pie chart
data2 = [
    ["Category", "Value"],
    ["Category 1", 30],
    ["Category 2", 45],
    ["Category 3", 25],
]

# Add the data to the sheet starting from a specific row
start_row = 38
for i, row in enumerate(data2, start=start_row):
    for j, value in enumerate(row, start=12):  # Start from column L (12th column)
        sheet.cell(row=i, column=j, value=value)

# Create a second pie chart
pie2 = PieChart()
labels = Reference(sheet, min_col=12, min_row=start_row + 1, max_row=start_row + 3)
data = Reference(sheet, min_col=13, min_row=start_row + 1, max_row=start_row + 3)
pie2.add_data(data, titles_from_data=True)
pie2.set_categories(labels)
pie2.title = "Sample Pie Chart 2"

# Set the size of the pie chart
pie2.width = 10  # Width in columns (L to R)
pie2.height = 12  # Height in rows (38 to 56)

# Add the pie chart to the sheet
sheet.add_chart(pie2, "L38")

# Add sample data for the third pie chart
data3 = [
    ["Category", "Value"],
    ["Category 1", 30],
    ["Category 2", 45],
    ["Category 3", 25],
]

# Add the data to the sheet starting from a specific row
start_row = 38
for i, row in enumerate(data3, start=start_row):
    for j, value in enumerate(row, start=19):  # Start from column S (19th column)
        sheet.cell(row=i, column=j, value=value)

# Create a third pie chart
pie3 = PieChart()
labels = Reference(sheet, min_col=19, min_row=start_row + 1, max_row=start_row + 3)
data = Reference(sheet, min_col=20, min_row=start_row + 1, max_row=start_row + 3)
pie3.add_data(data, titles_from_data=True)
pie3.set_categories(labels)
pie3.title = "Sample Pie Chart 3"

# Set the size of the pie chart
pie3.width = 10  # Width in columns (S to Y)
pie3.height = 12  # Height in rows (38 to 56)

# Add the pie chart to the sheet
sheet.add_chart(pie3, "S38")

# Save the workbook to a file named 'test.xlsx' in the same directory
workbook.save("test.xlsx")