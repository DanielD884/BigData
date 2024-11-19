import argparse
import ast
from functools import reduce
from os import path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
import pyspark
from pyspark.sql import DataFrame, SparkSession


def get_args():
    parser = argparse.ArgumentParser(
        description="Some Basic Spark Job doing some stuff on hubway data stored within HDFS."
    )
    parser.add_argument("--yearmonth", help="Partion Year Month", required=True, type=str)

    return parser.parse_args()


if __name__ == "__main__":
    """
    Main Function
    """

    # Parse Command Line Args
    args = get_args()
    year_months = ast.literal_eval(args.yearmonth)

    print(year_months)

    # Initialize Spark Context
    sc = pyspark.SparkContext()
    spark = SparkSession(sc)

    with pd.ExcelWriter("/home/airflow/output/combined-kpis.xlsx") as writer:
            # Loop over all year months
            for year_month in year_months:

                kpi_file = path.join("/user/hadoop/hubway_data/kpis", year_month, "kpis.parquet")

                # Load the current KPI file as a DataFrame
                df = (
                    spark.read.format("parquet")
                    .options(header="true", delimiter=",", nullValue="null", inferschema="true")
                    .load(kpi_file)
                )

                # Convert the Spark DataFrame to Pandas DataFrame and write to a specific sheet
                df.toPandas().to_excel(writer, sheet_name=year_month, index=False, header=True)
    
   # load the workbook
    workbook = load_workbook("/home/airflow/output/combined-kpis.xlsx")
    
    for year_month in year_months:
        sheet = workbook[year_month]

        # Remove gridlines
        sheet.sheet_view.showGridLines = False

        # Define the thick border style
        thick_border = Border(left=Side(style='thick', color='000000'),
                            right=Side(style='thick', color='000000'),
                            top=Side(style='thick', color='000000'),
                            bottom=Side(style='thick', color='000000'))

        # Remove all existing borders in the range D5 to Y62
        for row in range(5, 63):
            for col in range(4, 29):  
                cell = sheet.cell(row=row, column=col)
                cell.border = Border()

        # Apply the thick border around the range D5 to Y62
        for row in range(5, 63):
            # Left border
            sheet.cell(row=row, column=4).border = Border(left=thick_border.left)
            # Right border
            sheet.cell(row=row, column=29).border = Border(right=thick_border.right)

        for col in range(4, 29):
            # Top border
            sheet.cell(row=5, column=col).border = Border(top=thick_border.top)
            # Bottom border
            sheet.cell(row=62, column=col).border = Border(bottom=thick_border.bottom)

        # Set the corners
        sheet.cell(row=5, column=4).border = Border(top=thick_border.top, left=thick_border.left)
        sheet.cell(row=5, column=25).border = Border(top=thick_border.top, right=thick_border.right)
        sheet.cell(row=62, column=4).border = Border(bottom=thick_border.bottom, left=thick_border.left)
        sheet.cell(row=62, column=29).border = Border(bottom=thick_border.bottom, right=thick_border.right)

        # Set the text "hubway-data" in a merged cell
        sheet.merge_cells('D5:AB6')
        cell = sheet.cell(row=5, column=4)
        cell.value = "Hubway Data Dashboard"
        cell.font = Font(color="FFFFFF")  # White font
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Gray background
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

        # Merge cells E8 to H13 and set the month and year
        sheet.merge_cells('E8:H13')
        cell = sheet.cell(row=8, column=5)
        cell.value = sheet["A2"].value  
        cell.font = Font(color="000000")  # Black font
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

        # Apply thick border 
        for row in range(8, 14):
            sheet.cell(row=row, column=5).border = Border(left=thick_border.left)
            sheet.cell(row=row, column=8).border = Border(right=thick_border.right)

        for col in range(5, 9):
            sheet.cell(row=8, column=col).border = Border(top=thick_border.top)
            sheet.cell(row=13, column=col).border = Border(bottom=thick_border.bottom)

        # Set the corners 
        sheet.cell(row=8, column=5).border = Border(top=thick_border.top, left=thick_border.left)
        sheet.cell(row=8, column=8).border = Border(top=thick_border.top, right=thick_border.right)
        sheet.cell(row=13, column=5).border = Border(bottom=thick_border.bottom, left=thick_border.left)
        sheet.cell(row=13, column=8).border = Border(bottom=thick_border.bottom, right=thick_border.right)

        # Styleelements
        header_font = Font(name="Calibri", size=14, bold=True)
        value_font = Font(name="Calibri", size=12)
        header_fill = PatternFill(start_color="BCE4F7", end_color="BCE4F7", fill_type="solid")
        value_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        thick_border = Border(
            left=Side(style="thick"),
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick"),
        )

        # **"Average Trip Duration" Box**
        sheet.merge_cells('E15:H18')
        header_cell = sheet.cell(row=15, column=5, value="Average Trip Duration in Minutes")
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(15, 19):
            for col in range(5, 9):
                sheet.cell(row=row, column=col).fill = header_fill
                if row == 15 or row == 18 or col == 5 or col == 8:
                    sheet.cell(row=row, column=col).border = thick_border

        average_trip_duration = sheet["B2"].value  # Beispiel: Wert aus Zelle B2
        sheet.merge_cells('E19:H24')
        value_cell = sheet.cell(row=19, column=5, value=average_trip_duration)
        value_cell.font = value_font
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(19, 25):
            for col in range(5, 9):
                sheet.cell(row=row, column=col).fill = value_fill
                if row == 19 or row == 24 or col == 5 or col == 8:
                    sheet.cell(row=row, column=col).border = thick_border

        # **"Average Distance" Box**
        sheet.merge_cells('E26:H29')
        header_cell = sheet.cell(row=26, column=5, value="Average Distance in Kilometers")
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(26, 30):
            for col in range(5, 9):
                sheet.cell(row=row, column=col).fill = header_fill
                if row == 26 or row == 29 or col == 5 or col == 8:
                    sheet.cell(row=row, column=col).border = thick_border

        average_distance = sheet["C2"].value  
        sheet.merge_cells('E30:H35')
        value_cell = sheet.cell(row=30, column=5, value=average_distance)
        value_cell.font = value_font
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(30, 36):
            for col in range(5, 9):
                sheet.cell(row=row, column=col).fill = value_fill
                if row == 30 or row == 35 or col == 5 or col == 8:
                    sheet.cell(row=row, column=col).border = thick_border


        # Add the first table from K15 to N37 with columns "Bike ID" and "Number of Usage"
        sheet.merge_cells('K11:N12')
        cell = sheet.cell(row=11, column=11, value="Bike ID and Number of Usage")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sheet.merge_cells('K15:L16')
        sheet.merge_cells('M15:N16')
        cell = sheet.cell(row=15, column=11, value="Bike ID")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell = sheet.cell(row=15, column=13, value="Number of Usage")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Fill the first table from N2 to Z2 with data about the top 10 bikes
        for i in range(10):
            platz_cell = sheet.cell(row=17 + 2 * i, column=10, value=f"No. {i + 1}")
            platz_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.merge_cells(start_row=17 + 2 * i, start_column=10, end_row=18 + 2 * i, end_column=10)

            bike_id = sheet.cell(row=2, column=14 + 2 * i).value
            usage_count = sheet.cell(row=2, column=15 + 2 * i).value
            row_start = 17 + 2 * i
            row_end = row_start + 1
            sheet.merge_cells(start_row=row_start, start_column=11, end_row=row_end, end_column=12)
            sheet.merge_cells(start_row=row_start, start_column=13, end_row=row_end, end_column=14)
            sheet.cell(row=row_start, column=11, value=bike_id).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.cell(row=row_start, column=13, value=usage_count).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Apply thick border around the first table
        for row in range(17, 37):
            for col in range(10, 15):
                cell = sheet.cell(row=row, column=col)
                cell.border = thick_border

        # Add the second table from Q15 to T37 with columns "Top 10 most start stations" and "Number of Usage"
        sheet.merge_cells('Q11:T12')
        cell = sheet.cell(row=11, column=17, value="Top 10 most start stations and Number of Usage")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sheet.merge_cells('Q15:R16')
        sheet.merge_cells('S15:T16')
        cell = sheet.cell(row=15, column=17, value="Top 10 most start stations")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell = sheet.cell(row=15, column=19, value="Number of Usage")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Fill the second table with the data
        for i in range(10):
            platz_cell = sheet.cell(row=17 + 2 * i, column=16, value=f"Platz {i + 1}")
            platz_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.merge_cells(start_row=17 + 2 * i, start_column=16, end_row=18 + 2 * i, end_column=16)

            start_station = sheet.cell(row=2, column=34 + 2 * i).value
            usage_count = sheet.cell(row=2, column=35 + 2 * i).value
            row_start = 17 + 2 * i
            row_end = row_start + 1
            sheet.merge_cells(start_row=row_start, start_column=17, end_row=row_end, end_column=18)
            sheet.merge_cells(start_row=row_start, start_column=19, end_row=row_end, end_column=20)
            sheet.cell(row=row_start, column=17, value=start_station).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.cell(row=row_start, column=19, value=usage_count).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in range(17, 37):
            for col in range(16, 21):
                cell = sheet.cell(row=row, column=col)
                cell.border = thick_border

        # Add the third table from W15 to Z37 with columns "Top 10 most end stations" and "Number of Usage"
        sheet.merge_cells('W11:Z12')
        cell = sheet.cell(row=11, column=23, value="Top 10 most end stations and Number of Usage")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sheet.merge_cells('W15:X16')
        sheet.merge_cells('Y15:Z16')
        cell = sheet.cell(row=15, column=23, value="Top 10 most end stations")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell = sheet.cell(row=15, column=25, value="Number of Usage")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Fill the third table with data
        for i in range(10):
            platz_cell = sheet.cell(row=17 + 2 * i, column=22, value=f"No. {i + 1}")
            platz_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.merge_cells(start_row=17 + 2 * i, start_column=22, end_row=18 + 2 * i, end_column=22)

            station_name = sheet.cell(row=2, column=54 + 2 * i).value
            usage_count = sheet.cell(row=2, column=55 + 2 * i).value
            row_start = 17 + 2 * i
            row_end = row_start + 1
            sheet.merge_cells(start_row=row_start, start_column=23, end_row=row_end, end_column=24)
            sheet.merge_cells(start_row=row_start, start_column=25, end_row=row_end, end_column=26)
            sheet.cell(row=row_start, column=23, value=station_name).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.cell(row=row_start, column=25, value=usage_count).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Apply thick border around the third table
        for row in range(17, 37):
            for col in range(22, 27):
                cell = sheet.cell(row=row, column=col)
                cell.border = thick_border

        # ** First Bar Chart **
        bar_chart1 = BarChart()
        labels = ["Male", "Female", "Unknown"]
        for col, label in enumerate(labels, start=4):  
            sheet.cell(row=1, column=col, value=label)  

        data_ref = Reference(sheet, min_col=4, min_row=1, max_col=6, max_row=2)   


        bar_chart1.add_data(data_ref, titles_from_data=True)
        bar_chart1.title = "Usage Share by Gender"

        bar_chart1.width = 10
        bar_chart1.height = 12

        sheet.add_chart(bar_chart1, "E38")

        # Second Bar Chart
        bar_chart2 = BarChart()

        data_ref2 = Reference(sheet, min_col=7, min_row=1, max_col=13, max_row=2)  

        bar_chart2.add_data(data_ref2, titles_from_data=True)

        bar_chart2.title = "Usage Share by Age"
        bar_chart2.width = 10
        bar_chart2.height = 12

        sheet.add_chart(bar_chart2, "N38")

        # Third Bar Chart
        bar_chart3 = BarChart()

        data_ref_bv_by = Reference(sheet, min_col=74, min_row=1, max_col=77, max_row=2)  

        bar_chart3.add_data(data_ref_bv_by, titles_from_data=True)

        bar_chart3.title = "Sample Bar Chart for BV to BY"
        bar_chart3.width = 10
        bar_chart3.height = 12

        sheet.add_chart(bar_chart3, "H50")



    # Speichern der formatieren Excel-Datei
    workbook.save("/home/airflow/output/combined-kpis.xlsx")