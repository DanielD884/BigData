import argparse
import ast
from functools import reduce
from os import path
import pyspark
from pyspark.sql import DataFrame, SparkSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference


def get_args():
    parser = argparse.ArgumentParser(
        description="Combine KPI data from multiple year-month partitions stored in HDFS."
    )
    parser.add_argument("--yearmonth", help="Partition Year Month", required=True, type=str)
    return parser.parse_args()


def main():
    args = get_args()
    year_months = ast.literal_eval(args.yearmonth)

    print(year_months)

    # Initialize Spark Context
    sc = pyspark.SparkContext()
    spark = SparkSession(sc)

    data = []

    # Loop over all year months
    for year_month in year_months:
      
        kpi_file = path.join("/user/hadoop/hubway_data/kpis", year_month, "kpis.parquet")
        df = spark.read.format("parquet").load(kpi_file)
        data.append(df)

    if data:
        # Combine all dataframes into one
        combinated_data = reduce(DataFrame.union, data)

        rows = combinated_data.collect()
        workbook = Workbook()

        workbook.remove(workbook.active)

       
        for row in rows:
            year_month = str(row["year_month"])  
            
            if year_month not in workbook.sheetnames:
                workbook.create_sheet(title=year_month)

            sheet = workbook[year_month]

            # Remove gridlines
            sheet.sheet_view.showGridLines = False

            # Define the thick border style
            thick_border = Border(left=Side(style='thick', color='000000'),
                                  right=Side(style='thick', color='000000'),
                                  top=Side(style='thick', color='000000'),
                                  bottom=Side(style='thick', color='000000'))

            # Remove all existing borders in the range D5 to Y62
            for row in range(5, 63):
                for col in range(4, 26):  # D is the 4th column, Y is the 25th column
                    cell = sheet.cell(row=row, column=col)
                    cell.border = Border()

            # Apply the thick border around the range D5 to Y62
            for row in range(5, 63):
                # Left border
                sheet.cell(row=row, column=4).border = Border(left=thick_border.left)
                # Right border
                sheet.cell(row=row, column=25).border = Border(right=thick_border.right)

            for col in range(4, 26):
                # Top border
                sheet.cell(row=5, column=col).border = Border(top=thick_border.top)
                # Bottom border
                sheet.cell(row=62, column=col).border = Border(bottom=thick_border.bottom)

            # Set the corners
            sheet.cell(row=5, column=4).border = Border(top=thick_border.top, left=thick_border.left)
            sheet.cell(row=5, column=25).border = Border(top=thick_border.top, right=thick_border.right)
            sheet.cell(row=62, column=4).border = Border(bottom=thick_border.bottom, left=thick_border.left)
            sheet.cell(row=62, column=25).border = Border(bottom=thick_border.bottom, right=thick_border.right)

            # Set the text "hubway-data" in a merged cell
            sheet.merge_cells('D5:Y6')
            cell = sheet.cell(row=5, column=4)
            cell.value = "hubway-data"
            cell.font = Font(color="FFFFFF")  # White font
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Gray background
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

            # change layout of the year_month
            year = year_month[:4]
            month = year_month[4:]
            year_month = f"Monat: {month} Jahr: {year}"
            # Merge cells E8 to H13 and set the text "February"
            sheet.merge_cells('E8:H13')
            cell = sheet.cell(row=8, column=5)
            cell.value = year_month
            cell.font = Font(color="000000")  # Black font
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

            # Apply thick border around the "February" box
            for row in range(8, 14):
                # Left border
                sheet.cell(row=row, column=5).border = Border(left=thick_border.left)
                # Right border
                sheet.cell(row=row, column=8).border = Border(right=thick_border.right)

            for col in range(5, 9):
                # Top border
                sheet.cell(row=8, column=col).border = Border(top=thick_border.top)
                # Bottom border
                sheet.cell(row=13, column=col).border = Border(bottom=thick_border.bottom)

            # Set the corners for the "February" box
            sheet.cell(row=8, column=5).border = Border(top=thick_border.top, left=thick_border.left)
            sheet.cell(row=8, column=8).border = Border(top=thick_border.top, right=thick_border.right)
            sheet.cell(row=13, column=5).border = Border(bottom=thick_border.bottom, left=thick_border.left)
            sheet.cell(row=13, column=8).border = Border(bottom=thick_border.bottom, right=thick_border.right)

            # Add the "Average Trip Duration" header
            sheet.merge_cells('E15:H18')
            cell = sheet.cell(row=15, column=5, value="Average Trip Duration")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Apply thick border around the "Average Trip Duration" box
            for row in range(15, 19):
                # Left border
                sheet.cell(row=row, column=5).border = Border(left=thick_border.left)
                # Right border
                sheet.cell(row=row, column=8).border = Border(right=thick_border.right)

            for col in range(5, 9):
                # Top border
                sheet.cell(row=15, column=col).border = Border(top=thick_border.top)
                # Bottom border
                sheet.cell(row=18, column=col).border = Border(bottom=thick_border.bottom)

            # Set the corners for the "Average Trip Duration" box
            sheet.cell(row=15, column=5).border = Border(top=thick_border.top, left=thick_border.left)
            sheet.cell(row=15, column=8).border = Border(top=thick_border.top, right=thick_border.right)
            sheet.cell(row=18, column=5).border = Border(bottom=thick_border.bottom, left=thick_border.left)
            sheet.cell(row=18, column=8).border = Border(bottom=thick_border.bottom, right=thick_border.right)

            # Add sample data for average trip duration
            average_trip_duration = str(row["avg_trip_duration"]) + " min"
            sheet.merge_cells('E19:H24')
            cell = sheet.cell(row=19, column=5, value=average_trip_duration)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add the "Average Distance" header
            sheet.merge_cells('E26:H29')
            cell = sheet.cell(row=26, column=5, value="Average Distance")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Apply thick border around the "Average Distance" box
            for row in range(26, 30):
                # Left border
                sheet.cell(row=row, column=5).border = Border(left=thick_border.left)
                # Right border
                sheet.cell(row=row, column=8).border = Border(right=thick_border.right)

            for col in range(5, 9):
                # Top border
                sheet.cell(row=26, column=col).border = Border(top=thick_border.top)
                # Bottom border
                sheet.cell(row=29, column=col).border = Border(bottom=thick_border.bottom)

            # Set the corners for the "Average Distance" box
            sheet.cell(row=26, column=5).border = Border(top=thick_border.top, left=thick_border.left)
            sheet.cell(row=26, column=8).border = Border(top=thick_border.top, right=thick_border.right)
            sheet.cell(row=29, column=5).border = Border(bottom=thick_border.bottom, left=thick_border.left)
            sheet.cell(row=29, column=8).border = Border(bottom=thick_border.bottom, right=thick_border.right)

            # Add sample data for average distance
            average_distance = "5 km"
            sheet.merge_cells('E30:H35')
            cell = sheet.cell(row=30, column=5, value=average_distance)
            cell.alignment = Alignment(horizontal="center", vertical="center")




                            




    workbook.save("/home/airflow/output/combined-kpis.xlsx")
       
    # Stop Spark Context
    spark.stop()


if __name__ == "__main__":
    main()